import openpyxl
import os


path = 'C:\\Users\\kyeou\\Dropbox\\PC (2)\\Desktop\\Stuff On Github\\Scripts\\2020 transactions\\2020transactions.xlsx'

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

# column where totals are to be addded -> 11
# row -> 16
# set string of every row starting from 16 in column 11 with =sum(dA:dB)
# A equals the first number before the hypen in next column 12, and B equals the number after the hypen
# all of the numbers are 3 digits so they can be substringed all the same way
# the strings contained in the column 12 will be named RANGE
# A = RANGE[0:2], B = RANGE[4:6]



# add clause to see if transaction type is credit or debit and add that to their respective totals


for i in range(16, 77):
    condVar = sheet_obj.cell(row=i, column=12).value
    
    if (len(str(condVar)) >= 7):
        RANGE1 = (sheet_obj.cell(row=i, column=12))
        RANGE2 = (sheet_obj.cell(row=i, column=12))
        sRANGE1 = (RANGE1.value)[0:(RANGE1.value).find('-')]
        sRANGE2 = (RANGE2.value)[(RANGE1.value).find('-') + 1:]
        cS1 = "=SUM(D" + sRANGE1 + ":D" + sRANGE2 + ")"
        # print(cS1)
        sTS1 = sheet_obj.cell(row=i, column=11)
        sTS1.value = cS1
    else:
        RANGE3 = sheet_obj.cell(row=i, column=12)
        sRANGE3 = int(RANGE3.value)
        cS2 = "=SUM(D" + str(sRANGE3) + ")"
        # print(cS2)
        sTS2 = sheet_obj.cell(row=i, column=11)
        sTS2.value = cS2

wb_obj.save('C:\\Users\\kyeou\\Dropbox\\PC (2)\\Desktop\\Stuff On Github\\Scripts\\2020 transactions\\2020transactionsPY.xlsx')